"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  📸 노바소닉 아이케어 — 블로그 이미지/동영상 수집 도구
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  기능:
    1단계: 네이버 블로그 검색 API로 관련 글 목록 수집
    2단계: 각 블로그 본문 분석 (체험단/협찬 감지 + 이미지/동영상 추출)
    3단계: 파일 다운로드 및 폴더별 정리
    4단계: 엑셀 보고서 자동 생성

  사용법:
    python blog_crawler.py              → 전체 실행 (검색 + 다운로드)
    python blog_crawler.py --search     → 검색만 (목록 확인)
    python blog_crawler.py --download   → 이전 검색 결과로 다운로드만

  필요 패키지:
    pip install requests beautifulsoup4 lxml openpyxl python-dotenv
"""
import os, sys, json, time, re, argparse
import urllib.request, urllib.parse
from datetime import datetime
from pathlib import Path

print("[INFO] Loading modules...")
print(flush=True)

try:
    import requests
    print("  - requests OK")
except ImportError:
    print("\n[ERROR] 'requests' package not found.")
    print("  Run: pip install requests")
    input("\nPress Enter to exit...")
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
    print("  - beautifulsoup4 OK")
except ImportError:
    print("\n[ERROR] 'beautifulsoup4' package not found.")
    print("  Run: pip install beautifulsoup4 lxml")
    input("\nPress Enter to exit...")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    print("  - openpyxl OK")
except ImportError:
    print("\n[ERROR] 'openpyxl' package not found.")
    print("  Run: pip install openpyxl")
    input("\nPress Enter to exit...")
    sys.exit(1)

print("  - All modules loaded!")
print(flush=True)

# ─── 설정 ────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

try:
    from dotenv import load_dotenv
    env_path = os.path.join(SCRIPT_DIR, ".env")
    load_dotenv(env_path)
    print(f"[INFO] .env loaded from: {env_path}")
except ImportError:
    print("[WARNING] python-dotenv not available, reading from system env")

NAVER_CLIENT_ID = os.environ.get("NAVER_CLIENT_ID", "")
NAVER_CLIENT_SECRET = os.environ.get("NAVER_CLIENT_SECRET", "")

print(f"[INFO] NAVER_CLIENT_ID: {'SET (' + NAVER_CLIENT_ID[:6] + '...)' if NAVER_CLIENT_ID else 'NOT SET!'}")
print(f"[INFO] NAVER_CLIENT_SECRET: {'SET' if NAVER_CLIENT_SECRET else 'NOT SET!'}")
print(flush=True)

# 검색 설정
SEARCH_QUERIES = [
    "노바소닉 아이케어",
    "노바소닉 아이케어 후기",
    "노바소닉 아이케어 체험",
    "노바소닉 아이케어 리뷰",
    "노바소닉 안구세정기",
]

# ─── 출력 폴더 구조 ──────────────────────────────────────────
# blog_수집자료/
#   ├── _블로그_수집현황.xlsx     ← 엑셀 보고서
#   ├── _검색결과.json            ← 원본 검색 데이터
#   ├── 001_20240315_블로거A_제목/
#   │   ├── img_001.jpg
#   │   ├── img_002.jpg
#   │   └── _동영상URL.txt
#   ├── 002_20231120_블로거B_제목/
#   │   └── ...
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "blog_수집자료")
SEARCH_RESULT_FILE = os.path.join(OUTPUT_DIR, "_검색결과.json")
EXCEL_FILE = os.path.join(OUTPUT_DIR, "_블로그_수집현황.xlsx")

# 크롤링 설정
REQUEST_DELAY = 1.5
DOWNLOAD_DELAY = 0.5
REQUEST_TIMEOUT = 15

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
    "Referer": "https://blog.naver.com",
}

# ─── 체험단/협찬 감지 키워드 ──────────────────────────────────
SPONSORED_KEYWORDS = [
    "체험단", "체험", "협찬", "제공받", "원고료", "소정의",
    "업체로부터", "무상으로", "제품을 받", "지원받",
    "서포터즈", "리뷰어", "인플루언서", "광고",
    "대가를 받", "경제적 대가", "무료로 제공",
    "#체험단", "#협찬", "#광고", "#제공",
    "이 포스팅은", "이 글은", "본 포스팅은",
    "뷰티블로거", "파워블로거",
]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  1단계: 네이버 블로그 검색 API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def search_naver_blog(query, display=100, start=1, sort="sim"):
    encoded = urllib.parse.quote(query)
    url = f"https://openapi.naver.com/v1/search/blog.json?query={encoded}&display={display}&start={start}&sort={sort}"
    req = urllib.request.Request(url)
    req.add_header("X-Naver-Client-Id", NAVER_CLIENT_ID)
    req.add_header("X-Naver-Client-Secret", NAVER_CLIENT_SECRET)
    try:
        resp = urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT)
        return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"    ❌ API error: {e}")
        return None


def clean_html(text):
    return re.sub(r'<[^>]+>', '', text).strip()


def collect_blog_list():
    print("\n" + "=" * 60)
    print("  1/4  Naver Blog Search")
    print("=" * 60)

    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        print("\n  [ERROR] NAVER API keys not set!")
        print("     Check NAVER_CLIENT_ID / NAVER_CLIENT_SECRET in .env")
        return []

    all_results = {}
    for query in SEARCH_QUERIES:
        print(f"\n  Search: '{query}'")
        start = 1
        page_count = 0
        while start <= 1000:
            result = search_naver_blog(query, display=100, start=start, sort="sim")
            if not result or not result.get("items"):
                break
            items = result["items"]
            total = result.get("total", 0)
            page_count += 1
            for item in items:
                link = item.get("link", "")
                if "blog.naver.com" not in link and "m.blog.naver.com" not in link:
                    continue
                if link not in all_results:
                    postdate = item.get("postdate", "")
                    if postdate and "2021" <= postdate[:4] <= "2025":
                        all_results[link] = {
                            "title": clean_html(item.get("title", "")),
                            "description": clean_html(item.get("description", "")),
                            "blogger": item.get("bloggername", ""),
                            "bloggerlink": item.get("bloggerlink", ""),
                            "postdate": postdate,
                            "link": link,
                            "query": query,
                        }
            print(f"    Page {page_count}: {len(items)} results (total ~{total})")
            if len(items) < 100 or start + 100 > min(total, 1000):
                break
            start += 100
            time.sleep(0.3)

    blog_list = sorted(all_results.values(), key=lambda x: x.get("postdate", ""), reverse=True)

    year_stats = {}
    for b in blog_list:
        year = b["postdate"][:4]
        year_stats[year] = year_stats.get(year, 0) + 1

    print(f"\n  Found {len(blog_list)} blogs (2021~2025)")
    for year in sorted(year_stats.keys()):
        bar = "=" * min(year_stats[year], 30)
        print(f"    {year}: {year_stats[year]:3d} {bar}")

    print(f"\n  Blog list:")
    print("  " + "-" * 58)
    for i, b in enumerate(blog_list, 1):
        pd = b["postdate"]
        date_str = f"{pd[:4]}-{pd[4:6]}-{pd[6:8]}" if len(pd) == 8 else pd
        title_short = b["title"][:40] + "..." if len(b["title"]) > 40 else b["title"]
        print(f"  [{i:3d}] {date_str} | {b['blogger'][:12]:12s} | {title_short}")

    return blog_list


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  2단계: 블로그 본문 분석 (체험단 감지 + 이미지/동영상 추출)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def get_blog_id_and_logno(url):
    m = re.search(r'blog\.naver\.com/([^/\?]+)/(\d+)', url)
    if m:
        return m.group(1), m.group(2)
    m_id = re.search(r'blogId=([^&]+)', url)
    m_no = re.search(r'logNo=(\d+)', url)
    if m_id and m_no:
        return m_id.group(1), m_no.group(1)
    m = re.search(r'm\.blog\.naver\.com/([^/\?]+)/(\d+)', url)
    if m:
        return m.group(1), m.group(2)
    return None, None


def get_postview_url(blog_id, log_no):
    return f"https://blog.naver.com/PostView.naver?blogId={blog_id}&logNo={log_no}&direct=true"


def to_original_url(src):
    """네이버 이미지 URL → 원본 최대 사이즈 URL 목록 (우선순위순)

    네이버 CDN은 ?type= 파라미터 없이는 403을 반환하는 경우가 많음.
    따라서 파라미터를 제거하는 대신, 큰 사이즈로 교체하는 전략 사용.
    여러 URL 후보를 반환하여 다운로드 시 순차 시도.
    """
    if not src:
        return []

    # 프로토콜 보정
    if src.startswith("//"):
        src = "https:" + src

    # 기본 URL (쿼리 파라미터 제거한 원본 경로)
    base_url = re.sub(r'\?.*$', '', src)
    # 썸네일 경로 보정
    base_url = re.sub(r'/w\d+_blur', '', base_url)
    base_url = re.sub(r'/w\d+_default', '', base_url)

    # 여러 URL 후보 생성 (다운로드 시 순차 시도)
    candidates = []

    # 1순위: ?type=w2 (네이버 원본 사이즈 — 가장 큰 이미지)
    candidates.append(base_url + "?type=w2")

    # 2순위: ?type=w966 (일반적으로 가장 잘 동작하는 큰 사이즈)
    candidates.append(base_url + "?type=w966")

    # 3순위: 파라미터 없는 원본 URL
    candidates.append(base_url)

    # 4순위: blogpfs 서버로 변환 (원본 서버)
    if "blogpfthumb-phinf" in base_url:
        pfs_url = base_url.replace("blogpfthumb-phinf", "blogpfs-phinf")
        candidates.append(pfs_url + "?type=w2")
        candidates.append(pfs_url)

    # 5순위: thumbnail → original 경로 변환
    if "/thumbnail/" in base_url:
        orig_url = base_url.replace("/thumbnail/", "/original/")
        candidates.append(orig_url)

    return candidates


def detect_sponsored(text):
    """체험단/협찬 키워드 감지 → (is_sponsored, matched_keywords)"""
    text_lower = text.lower()
    matched = []
    for kw in SPONSORED_KEYWORDS:
        if kw in text_lower or kw in text:
            matched.append(kw)
    return len(matched) > 0, matched


def analyze_blog(blog_url):
    """블로그 본문 분석: 체험단 감지 + 이미지 + 동영상 추출

    Returns: (is_sponsored, sponsor_keywords, body_text, images, videos)
    """
    blog_id, log_no = get_blog_id_and_logno(blog_url)
    if not blog_id or not log_no:
        return False, [], "", [], []

    postview_url = get_postview_url(blog_id, log_no)
    try:
        resp = requests.get(postview_url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")
    except Exception as e:
        print(f"    Page load failed: {e}")
        return False, [], "", [], []

    # ── 본문 텍스트 추출 (체험단 감지용) ──
    body_text = ""
    # SmartEditor ONE
    for se in soup.select(".se-main-container"):
        body_text += se.get_text(separator=" ", strip=True) + " "
    # 구형 에디터
    if not body_text.strip():
        for area in soup.select("#postViewArea, .post-view, .se_textarea"):
            body_text += area.get_text(separator=" ", strip=True) + " "
    # 최종 fallback
    if not body_text.strip():
        body_text = soup.get_text(separator=" ", strip=True)

    # ── 체험단/협찬 감지 ──
    is_sponsored, sponsor_kws = detect_sponsored(body_text)

    # ── 이미지 추출 ──
    # 각 이미지는 { "candidates": [url1, url2, ...] } 형태로 저장
    # 다운로드 시 후보 URL을 순차적으로 시도
    raw_srcs = []  # 중복 체크용 (기본URL 기준)
    images = []    # [ [candidate_url1, candidate_url2, ...], ... ]

    def _add_image(src):
        """이미지 URL을 후보 목록으로 변환하여 추가 (중복 제거)"""
        if not src or "pstatic.net" not in src:
            return
        base = re.sub(r'\?.*$', '', src)
        if base.startswith("//"):
            base = "https:" + base
        # 이미 추가된 이미지인지 기본URL로 확인
        base_clean = re.sub(r'/w\d+_(blur|default)', '', base)
        if base_clean in raw_srcs:
            return
        raw_srcs.append(base_clean)
        candidates = to_original_url(src)
        if candidates:
            images.append(candidates)

    # 1) se-image (스마트에디터 ONE) — 모든 속성에서 추출
    for img in soup.select(".se-image-resource"):
        for attr in ["data-origin-src", "data-lazy-src", "src"]:
            src = img.get(attr, "")
            if src and "pstatic.net" in src:
                _add_image(src)
                break  # 한 img에서 하나만

    # 2) se-module-image 내 a 태그 href
    for a_tag in soup.select(".se-module-image a"):
        _add_image(a_tag.get("href", ""))

    # 3) 일반 img 태그 (구형 에디터 + 모든 에디터)
    for img in soup.find_all("img"):
        for attr in ["data-origin-src", "data-lazy-src", "data-src", "src"]:
            src = img.get(attr, "")
            if src and ("blogfiles.pstatic.net" in src or "postfiles.pstatic.net" in src
                         or "blogpfthumb-phinf" in src or "blogpfs-phinf" in src
                         or "post-phinf.pstatic.net" in src):
                _add_image(src)
                break

    # 4) 스타일 배경 이미지 (일부 블로그에서 사용)
    for div in soup.select("[style*='pstatic.net']"):
        style = div.get("style", "")
        urls = re.findall(r'url\(["\']?(https?://[^"\')\s]+pstatic\.net[^"\')\s]*)["\']?\)', style)
        for url in urls:
            _add_image(url)

    # 5) og:image 메타태그 (최소한의 대표 이미지 — fallback)
    if not images:
        og_img = soup.find("meta", property="og:image")
        if og_img and og_img.get("content"):
            _add_image(og_img["content"])

    # ── 동영상 추출 ──
    videos = []
    for vid in soup.select(".se-video-resource"):
        vid_src = vid.get("data-src") or vid.get("src") or ""
        if vid_src:
            videos.append({"type": "naver_video", "url": vid_src})
    for iframe in soup.find_all("iframe"):
        iframe_src = iframe.get("src") or ""
        if "youtube.com" in iframe_src or "youtu.be" in iframe_src:
            videos.append({"type": "youtube", "url": iframe_src})
        elif "tv.naver.com" in iframe_src or "serviceapi.nmv" in iframe_src:
            videos.append({"type": "naver_tv", "url": iframe_src})
    for video_tag in soup.find_all("video"):
        vid_src = video_tag.get("src") or ""
        source_tag = video_tag.find("source")
        if source_tag:
            vid_src = source_tag.get("src") or vid_src
        if vid_src:
            videos.append({"type": "direct_video", "url": vid_src})
    for module in soup.select("[data-module='movie']"):
        movie_data = module.get("data-module-data") or ""
        if movie_data:
            try:
                mdata = json.loads(movie_data)
                vid_url = mdata.get("videoUrl") or mdata.get("url") or ""
                if vid_url:
                    videos.append({"type": "naver_blog_video", "url": vid_url})
            except:
                pass

    # 중복 제거 (images는 이미 _add_image에서 중복 제거됨)
    seen = set()
    unique_vids = []
    for v in videos:
        if v["url"] not in seen:
            seen.add(v["url"])
            unique_vids.append(v)

    return is_sponsored, sponsor_kws, body_text[:500], images, unique_vids


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  3단계: 파일 다운로드
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def safe_filename(name, max_len=50):
    name = re.sub(r'[\\/:*?"<>|\r\n\t]', '', name)
    name = name.strip(". ")
    return (name[:max_len] if len(name) > max_len else name) or "untitled"


def _format_size(size):
    if size > 1024 * 1024:
        return f"{size / 1024 / 1024:.1f}MB"
    elif size > 1024:
        return f"{size / 1024:.0f}KB"
    return f"{size}B"


def download_single_url(url, save_path):
    """단일 URL 다운로드 시도. 성공 시 (True, filesize) 반환."""
    try:
        dl_headers = HEADERS.copy()
        dl_headers["Accept"] = "image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8"
        resp = requests.get(url, headers=dl_headers, timeout=30, stream=True)
        resp.raise_for_status()

        # Content-Type 확인 — HTML이 오면 이미지가 아님
        ctype = resp.headers.get("Content-Type", "")
        if "text/html" in ctype:
            return False, 0

        total_size = 0
        with open(save_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)
                total_size += len(chunk)

        return True, total_size
    except Exception:
        return False, 0


def download_image_with_fallback(candidate_urls, save_path, min_size=10000):
    """후보 URL 목록을 순차 시도하여 가장 큰 이미지를 다운로드.

    Args:
        candidate_urls: [url1, url2, ...] 우선순위순
        save_path: 저장 경로
        min_size: 최소 파일 크기 (이하면 다음 후보 시도)

    Returns: (success, final_size)
    """
    best_size = 0
    best_url = ""

    for url in candidate_urls:
        ok, fsize = download_single_url(url, save_path)
        if ok and fsize > best_size:
            best_size = fsize
            best_url = url
            # 100KB 이상이면 충분히 큰 원본 — 더 시도 안 함
            if fsize >= 100_000:
                break
        time.sleep(0.2)

    # 최고 결과가 min_size 이하면 실패
    if best_size < min_size:
        if os.path.exists(save_path):
            os.remove(save_path)
        return False, 0

    # 최고 URL로 최종 다운로드 (마지막 시도가 best가 아닐 수 있으므로)
    if best_url and best_url != candidate_urls[-1]:
        download_single_url(best_url, save_path)

    return True, best_size


def get_extension_from_url(url):
    if isinstance(url, list):
        url = url[0] if url else ""
    path = urllib.parse.urlparse(url).path
    ext = os.path.splitext(path)[1].lower()
    return ext if ext in [".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".svg"] else ".jpg"


def download_blog_media(blog_info, images, videos, base_dir, folder_num):
    """한 블로그 이미지/동영상 다운로드 → 개별 폴더

    images: [ [candidate_url1, url2, ...], [candidate_url1, url2, ...], ... ]
    """
    pd = blog_info["postdate"]
    date_str = f"{pd[:4]}{pd[4:6]}{pd[6:8]}" if len(pd) == 8 else pd
    folder_name = safe_filename(
        f"{folder_num:03d}_{date_str}_{blog_info['blogger']}_{blog_info['title']}", max_len=80
    )
    blog_dir = os.path.join(base_dir, folder_name)

    if not images and not videos:
        return 0, 0, ""

    os.makedirs(blog_dir, exist_ok=True)

    # ── 이미지 다운로드 (후보 URL 순차 시도) ──
    img_count = 0
    if images:
        for i, candidate_urls in enumerate(images, 1):
            ext = get_extension_from_url(candidate_urls)
            filename = f"img_{i:03d}{ext}"
            save_path = os.path.join(blog_dir, filename)

            print(f"      [{i}/{len(images)}] Trying {len(candidate_urls)} URLs... ", end="", flush=True)
            ok, fsize = download_image_with_fallback(candidate_urls, save_path, min_size=5000)

            if ok:
                print(f"OK ({_format_size(fsize)})")
                img_count += 1
            else:
                print(f"SKIP (too small or failed)")
                if os.path.exists(save_path):
                    os.remove(save_path)

            time.sleep(DOWNLOAD_DELAY)

    # ── 동영상 URL 저장 ──
    vid_count = 0
    if videos:
        vid_list_path = os.path.join(blog_dir, "_동영상_URL목록.txt")
        with open(vid_list_path, "w", encoding="utf-8") as f:
            f.write(f"# {blog_info['title']}\n# {blog_info['link']}\n\n")
            for i, vid in enumerate(videos, 1):
                f.write(f"[{i}] {vid['type']}: {vid['url']}\n")
        vid_count = len(videos)

        for i, vid in enumerate(videos, 1):
            if vid["type"] == "direct_video" and vid["url"]:
                vid_url = vid["url"]
                if vid_url.startswith("//"):
                    vid_url = "https:" + vid_url
                filename = f"video_{i:03d}.mp4"
                save_path = os.path.join(blog_dir, filename)
                print(f"      [vid {i}] ", end="", flush=True)
                ok, fsize = download_single_url(vid_url, save_path)
                if ok:
                    print(f"OK ({_format_size(fsize)})")
                else:
                    print("FAIL")
                time.sleep(DOWNLOAD_DELAY)

    return img_count, vid_count, blog_dir


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  4단계: 엑셀 보고서 생성
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def create_excel_report(report_data, excel_path):
    """수집 결과 엑셀 보고서 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "블로그 수집현황"

    # ── 스타일 ──
    header_fill = PatternFill("solid", fgColor="1B2A4A")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    sponsored_fill = PatternFill("solid", fgColor="FFF2CC")   # 연한 노랑 (체험단)
    normal_fill = PatternFill("solid", fgColor="FFFFFF")
    link_font = Font(name="맑은 고딕", color="2E6DB4", underline="single", size=10)
    data_font = Font(name="맑은 고딕", size=10)
    border = Border(
        left=Side("thin", "BBBBBB"), right=Side("thin", "BBBBBB"),
        top=Side("thin", "BBBBBB"), bottom=Side("thin", "BBBBBB"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── 헤더 ──
    headers = [
        ("No", 5),
        ("날짜", 12),
        ("블로거", 15),
        ("제목", 40),
        ("체험단/협찬", 13),
        ("감지 키워드", 25),
        ("이미지 수", 10),
        ("동영상 수", 10),
        ("이미지 폴더", 30),
        ("블로그 URL", 45),
    ]

    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── 데이터 ──
    for row_idx, data in enumerate(report_data, 2):
        pd_str = data["postdate"]
        date_str = f"{pd_str[:4]}-{pd_str[4:6]}-{pd_str[6:8]}" if len(pd_str) == 8 else pd_str
        is_sponsored = data.get("is_sponsored", False)
        row_fill = sponsored_fill if is_sponsored else normal_fill

        row_values = [
            row_idx - 1,                                     # No
            date_str,                                        # 날짜
            data["blogger"],                                 # 블로거
            data["title"],                                   # 제목
            "O 체험/협찬" if is_sponsored else "",           # 체험단/협찬
            ", ".join(data.get("sponsor_keywords", [])),     # 감지 키워드
            data.get("image_count", 0),                      # 이미지 수
            data.get("video_count", 0),                      # 동영상 수
            data.get("folder_name", ""),                     # 이미지 폴더
            data["link"],                                    # URL
        ]

        for col, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = center if col in [1, 2, 5, 7, 8] else left

            if col == 10:  # URL 링크
                cell.font = link_font
                cell.hyperlink = val
            elif col == 5 and is_sponsored:
                cell.font = Font(name="맑은 고딕", bold=True, color="C00000", size=10)
            else:
                cell.font = data_font

    # ── 요약 시트 ──
    ws2 = wb.create_sheet("요약")
    total = len(report_data)
    sponsored_count = sum(1 for d in report_data if d.get("is_sponsored"))
    total_imgs = sum(d.get("image_count", 0) for d in report_data)
    total_vids = sum(d.get("video_count", 0) for d in report_data)

    summary_data = [
        ["항목", "값"],
        ["수집 일시", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["총 블로그 수", total],
        ["체험단/협찬 블로그", f"{sponsored_count}개 ({sponsored_count/total*100:.0f}%)" if total > 0 else "0"],
        ["일반 블로그", total - sponsored_count],
        ["총 이미지 수", total_imgs],
        ["총 동영상 수", total_vids],
        ["검색어", ", ".join(SEARCH_QUERIES)],
    ]

    for r, (label, val) in enumerate(summary_data, 1):
        ws2.cell(row=r, column=1, value=label).font = Font(name="맑은 고딕", bold=(r == 1), size=11)
        ws2.cell(row=r, column=2, value=val).font = Font(name="맑은 고딕", size=11)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50

    # ── 연도별 시트 ──
    year_stats = {}
    for d in report_data:
        year = d["postdate"][:4]
        if year not in year_stats:
            year_stats[year] = {"total": 0, "sponsored": 0, "images": 0}
        year_stats[year]["total"] += 1
        if d.get("is_sponsored"):
            year_stats[year]["sponsored"] += 1
        year_stats[year]["images"] += d.get("image_count", 0)

    ws3 = wb.create_sheet("연도별 통계")
    ws3.append(["연도", "블로그 수", "체험단/협찬", "일반", "이미지 수"])
    for year in sorted(year_stats.keys()):
        s = year_stats[year]
        ws3.append([f"{year}년", s["total"], s["sponsored"], s["total"] - s["sponsored"], s["images"]])
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["E"].width = 12

    # 필터 설정
    ws.auto_filter.ref = ws.dimensions

    wb.save(excel_path)
    print(f"\n  Excel saved: {excel_path}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  메인 실행
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def run_search():
    blog_list = collect_blog_list()
    if not blog_list:
        print("\n  No results found.")
        return []
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(SEARCH_RESULT_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "search_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "queries": SEARCH_QUERIES,
            "total": len(blog_list),
            "blogs": blog_list,
        }, f, ensure_ascii=False, indent=2)
    print(f"\n  Search results saved: {SEARCH_RESULT_FILE}")
    return blog_list


def run_download(blog_list=None):
    if blog_list is None:
        if not os.path.exists(SEARCH_RESULT_FILE):
            print("\n  [ERROR] No search results. Run --search first.")
            return
        with open(SEARCH_RESULT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            blog_list = data.get("blogs", [])
    if not blog_list:
        print("\n  No blogs to download.")
        return

    print("\n" + "=" * 60)
    print("  2/4  Analyze blogs + Download media")
    print("=" * 60)
    print(f"  Target: {len(blog_list)} blogs\n")

    report_data = []

    for i, blog in enumerate(blog_list, 1):
        pd = blog["postdate"]
        date_str = f"{pd[:4]}-{pd[4:6]}-{pd[6:8]}" if len(pd) == 8 else pd
        title_short = blog["title"][:35] + "..." if len(blog["title"]) > 35 else blog["title"]

        print(f"  [{i}/{len(blog_list)}] {date_str} | {title_short}")
        print(f"    URL: {blog['link']}")

        # ── 본문 분석 ──
        is_sponsored, sponsor_kws, body_snippet, images, videos = analyze_blog(blog["link"])

        sponsor_tag = " [SPONSORED]" if is_sponsored else ""
        print(f"    -> Images: {len(images)} | Videos: {len(videos)}{sponsor_tag}")
        if sponsor_kws:
            print(f"    -> Keywords: {', '.join(sponsor_kws[:5])}")

        # ── 다운로드 ──
        img_count, vid_count, folder_path = 0, 0, ""
        if images or videos:
            img_count, vid_count, folder_path = download_blog_media(
                blog, images, videos, OUTPUT_DIR, i
            )

        # ── 보고서 데이터 ──
        report_data.append({
            **blog,
            "is_sponsored": is_sponsored,
            "sponsor_keywords": sponsor_kws,
            "image_count": img_count,
            "video_count": vid_count,
            "folder_name": os.path.basename(folder_path) if folder_path else "",
        })

        time.sleep(REQUEST_DELAY)

    # ── 결과 요약 ──
    total_imgs = sum(d["image_count"] for d in report_data)
    total_vids = sum(d["video_count"] for d in report_data)
    sponsored_count = sum(1 for d in report_data if d["is_sponsored"])
    success_count = sum(1 for d in report_data if d["image_count"] > 0 or d["video_count"] > 0)

    print("\n" + "=" * 60)
    print("  3/4  Summary")
    print("=" * 60)
    print(f"  Blogs analyzed : {len(blog_list)}")
    print(f"  Sponsored/Trial: {sponsored_count} blogs")
    print(f"  Media collected: {success_count} blogs")
    print(f"  Total images   : {total_imgs}")
    print(f"  Total videos   : {total_vids}")
    print(f"  Output folder  : {OUTPUT_DIR}")

    # ── 엑셀 생성 ──
    print("\n" + "=" * 60)
    print("  4/4  Creating Excel report")
    print("=" * 60)
    create_excel_report(report_data, EXCEL_FILE)

    # ── JSON 결과도 업데이트 ──
    with open(SEARCH_RESULT_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "search_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "queries": SEARCH_QUERIES,
            "total": len(report_data),
            "blogs": report_data,
        }, f, ensure_ascii=False, indent=2, default=str)

    print("\n  All done!")


def main():
    parser = argparse.ArgumentParser(description="Novasonic EyeCare Blog Crawler")
    parser.add_argument("--search", action="store_true", help="Search only")
    parser.add_argument("--download", action="store_true", help="Download from previous search")
    parser.add_argument("--query", nargs="+", help="Add search queries")
    args = parser.parse_args()

    print()
    print("=" * 60)
    print("  Novasonic EyeCare - Blog Media Collector")
    print("=" * 60)
    print(f"  Time: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    if args.query:
        SEARCH_QUERIES.extend(args.query)
        print(f"  Extra queries: {args.query}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if args.download:
        run_download()
    elif args.search:
        run_search()
    else:
        blog_list = run_search()
        if blog_list:
            print("\n" + "-" * 60)
            confirm = input("  Start download? (y/n): ").strip().lower()
            if confirm == 'y':
                run_download(blog_list)
            else:
                print("  Skipped. Run later: python blog_crawler.py --download")

    print()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  Cancelled by user.")
    except Exception as e:
        print(f"\n\n  [ERROR] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print()
        input("  Press Enter to exit...")
