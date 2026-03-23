"""
STEP 3 — 네이버 쇼핑 API로 최저가 자동 조회 → 엑셀 업데이트
"""
import os, json, time, re
import urllib.request, urllib.parse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import NAVER_CLIENT_ID, NAVER_CLIENT_SECRET, NAVER_DISPLAY

def _fill(c):
    return PatternFill("solid", start_color=c, fgColor=c)

def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── 네이버 쇼핑 검색 API ───────────────────────────────────
def search_naver_shopping(query, display=NAVER_DISPLAY):
    encoded = urllib.parse.quote(query)
    url = f"https://openapi.naver.com/v1/search/shop.json?query={encoded}&display={display}&sort=asc"

    req = urllib.request.Request(url)
    req.add_header("X-Naver-Client-Id", NAVER_CLIENT_ID)
    req.add_header("X-Naver-Client-Secret", NAVER_CLIENT_SECRET)

    try:
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read().decode("utf-8"))
        return data.get("items", [])
    except urllib.error.HTTPError as e:
        print(f"   ⚠️  HTTP 오류 {e.code}: {e.reason}")
        return []
    except Exception as e:
        print(f"   ⚠️  검색 오류: {e}")
        return []

def find_lowest_price(items):
    """검색 결과에서 최저가 항목 추출"""
    if not items:
        return None

    lowest = None
    for item in items:
        try:
            price = int(item.get("lprice", 0))
            if price > 0:
                if lowest is None or price < lowest["price"]:
                    # HTML 태그 제거
                    title = re.sub(r"<[^>]+>", "", item.get("title", ""))
                    lowest = {
                        "price": price,
                        "price_str": f"{price:,}원",
                        "mall": item.get("mallName", ""),
                        "title": title,
                        "link": item.get("link", ""),
                        "image": item.get("image", ""),
                    }
        except (ValueError, TypeError):
            continue

    return lowest

# ─── 엑셀 업데이트 ──────────────────────────────────────────
def update_excel(excel_path, product_list):
    wb = load_workbook(excel_path)
    ws = wb.active
    border = _border()

    results = []
    for product in product_list:
        row = product["row"]
        name = product["name"]
        brand = product.get("brand", "")

        # 검색어 조합: 브랜드 + 제품명
        query = f"{brand} {name}".strip() if brand else name
        print(f"   [{product_list.index(product)+1}/{len(product_list)}] 검색: {query}")

        items = search_naver_shopping(query)
        lowest = find_lowest_price(items)

        shade = "F5F5F5" if (product_list.index(product)) % 2 == 0 else "FFFFFF"

        # 네이버 최저가 컬럼: 20(최저가), 21(쇼핑몰), 22(비고)
        col_price = 20
        col_mall  = 21
        col_note  = 22

        if lowest:
            c_price = ws.cell(row=row, column=col_price)
            c_price.value = lowest["price_str"]
            c_price.font = Font(name="맑은 고딕", bold=True, size=10, color="C00000")
            c_price.fill, c_price.alignment, c_price.border = _fill(shade), Alignment(horizontal="center", vertical="center"), border

            c_mall = ws.cell(row=row, column=col_mall)
            c_mall.value = lowest["mall"]
            c_mall.font = Font(name="맑은 고딕", size=9, color="333333")
            c_mall.fill, c_mall.alignment, c_mall.border = _fill(shade), Alignment(horizontal="center", vertical="center"), border

            c_note = ws.cell(row=row, column=col_note)
            c_note.value = lowest["title"][:50]
            c_note.font = Font(name="맑은 고딕", size=8, color="888888")
            c_note.fill, c_note.alignment, c_note.border = _fill(shade), Alignment(horizontal="left", vertical="center", wrap_text=True), border

            print(f"      → 최저가: {lowest['price_str']} ({lowest['mall']})")
            results.append({"product_name": name, "price": lowest["price"], "mall_name": lowest["mall"], **lowest})
        else:
            c_price = ws.cell(row=row, column=col_price)
            c_price.value = "조회불가"
            c_price.font = Font(name="맑은 고딕", size=9, color="AAAAAA")
            c_price.fill, c_price.alignment, c_price.border = _fill(shade), Alignment(horizontal="center", vertical="center"), border

            c_note = ws.cell(row=row, column=col_note)
            search_note = f"{query} 검색결과 없음"
            c_note.value = search_note
            c_note.font = Font(name="맑은 고딕", size=8, color="AAAAAA")
            c_note.fill, c_note.alignment, c_note.border = _fill(shade), Alignment(horizontal="left", vertical="center", wrap_text=True), border

            print(f"      → 검색결과 없음")
            results.append({"product_name": name, "price": 0, "mall_name": ""})

        time.sleep(0.3)  # API 호출 간격

    wb.save(excel_path)
    print(f"\n✅ 엑셀 업데이트 완료 → {excel_path}")
    return results

# ─── 메인 실행 ─────────────────────────────────────────────
def run(product_list, excel_path):
    print("=" * 60)
    print("  STEP 3: 네이버 쇼핑 최저가 자동 조회")
    print("=" * 60)

    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        print("\n⚠️  네이버 API 키가 설정되지 않았습니다.")
        print("   config.py에서 NAVER_CLIENT_ID, NAVER_CLIENT_SECRET을 설정해 주세요.")
        return []

    print(f"\n🔍 {len(product_list)}개 제품 최저가를 조회합니다...\n")
    results = update_excel(excel_path, product_list)

    # 요약
    found = [r for r in results if r["price"] > 0]
    print(f"\n📊 조회 결과: {len(found)}/{len(results)}개 제품 최저가 확인")
    for r in found:
        print(f"   - {r['product']}: {r['price_str']} ({r['mall']})")

    return results

if __name__ == "__main__":
    print("이 모듈은 main.py를 통해 실행해 주세요.")
