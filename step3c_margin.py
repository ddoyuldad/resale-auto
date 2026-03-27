"""
STEP 3-3 — 마진율 자동 계산

매입가(매대 확인 가격)와 온라인 판매가를 비교하여 마진을 계산.
  - 네이버: 수수료 6%  →  순수익 = 판매가 × 0.94 - 매입가
  - 쿠팡:   수수료 12% →  순수익 = 판매가 × 0.88 - 매입가
  - 마진율 = 순수익 / 매입가 × 100
"""
import os, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVER_FEE  = 0.06   # 네이버 수수료 6%
COUPANG_FEE = 0.12  # 쿠팡 수수료 12%

def _fill(c):
    return PatternFill("solid", start_color=c, fgColor=c)

def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def parse_price(val):
    """가격 문자열에서 숫자 추출. '12,500원' → 12500"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(",", "").replace("원", "").replace(" ", "").strip()
    match = re.search(r"(\d+)", s)
    return int(match.group(1)) if match else 0


def calc_margin(buy_price, sell_price, fee_rate):
    """마진 계산. (마진액, 마진율%) 반환"""
    if buy_price <= 0 or sell_price <= 0:
        return 0, 0.0
    net = sell_price * (1 - fee_rate)
    margin = net - buy_price
    margin_pct = (margin / buy_price) * 100
    return int(margin), round(margin_pct, 1)


def run(excel_path, product_list):
    """엑셀에서 매입가/판매가를 읽어 마진 컬럼 자동 계산"""
    print("=" * 60)
    print("  STEP 3-3: 마진율 자동 계산")
    print("=" * 60)

    if not excel_path or not os.path.exists(excel_path):
        print(f"\n⚠️  엑셀 파일을 찾을 수 없습니다: {excel_path}")
        return []

    wb = load_workbook(excel_path)
    ws = wb.active
    border = _border()

    results = []

    for i, product in enumerate(product_list):
        row = product["row"]
        name = product["name"]

        # 공급가 읽기 (G열 = 7번 컬럼: "공급가")
        buy_price = parse_price(ws.cell(row=row, column=7).value)

        # 네이버 최저가 읽기 (T열 = 20번 컬럼)
        naver_price = parse_price(ws.cell(row=row, column=20).value)

        # 쿠팡 최저가 읽기 (W열 = 23번 컬럼)
        coupang_price = parse_price(ws.cell(row=row, column=23).value)

        shade = "F5F5F5" if i % 2 == 0 else "FFFFFF"

        # 네이버 마진 계산
        n_margin, n_pct = calc_margin(buy_price, naver_price, NAVER_FEE)
        # 쿠팡 마진 계산
        c_margin, c_pct = calc_margin(buy_price, coupang_price, COUPANG_FEE)

        # Y열(25): 네이버 마진
        cell_nm = ws.cell(row=row, column=25)
        if naver_price > 0 and buy_price > 0:
            cell_nm.value = f"{n_margin:,}원"
            color = "0D7C3F" if n_margin > 0 else "C00000"
            cell_nm.font = Font(name="맑은 고딕", bold=True, size=10, color=color)
        else:
            cell_nm.value = "-"
            cell_nm.font = Font(name="맑은 고딕", size=9, color="AAAAAA")
        cell_nm.fill = _fill(shade)
        cell_nm.alignment = Alignment(horizontal="center", vertical="center")
        cell_nm.border = border

        # Z열(26): 네이버 마진율
        cell_np = ws.cell(row=row, column=26)
        if naver_price > 0 and buy_price > 0:
            cell_np.value = f"{n_pct:+.1f}%"
            color = "0D7C3F" if n_pct > 0 else "C00000"
            cell_np.font = Font(name="맑은 고딕", bold=True, size=11, color=color)
        else:
            cell_np.value = "-"
            cell_np.font = Font(name="맑은 고딕", size=9, color="AAAAAA")
        cell_np.fill = _fill(shade)
        cell_np.alignment = Alignment(horizontal="center", vertical="center")
        cell_np.border = border

        # AA열(27): 쿠팡 마진
        cell_cm = ws.cell(row=row, column=27)
        if coupang_price > 0 and buy_price > 0:
            cell_cm.value = f"{c_margin:,}원"
            color = "0D7C3F" if c_margin > 0 else "C00000"
            cell_cm.font = Font(name="맑은 고딕", bold=True, size=10, color=color)
        else:
            cell_cm.value = "-"
            cell_cm.font = Font(name="맑은 고딕", size=9, color="AAAAAA")
        cell_cm.fill = _fill(shade)
        cell_cm.alignment = Alignment(horizontal="center", vertical="center")
        cell_cm.border = border

        # AB열(28): 쿠팡 마진율
        cell_cp = ws.cell(row=row, column=28)
        if coupang_price > 0 and buy_price > 0:
            cell_cp.value = f"{c_pct:+.1f}%"
            color = "0D7C3F" if c_pct > 0 else "C00000"
            cell_cp.font = Font(name="맑은 고딕", bold=True, size=11, color=color)
        else:
            cell_cp.value = "-"
            cell_cp.font = Font(name="맑은 고딕", size=9, color="AAAAAA")
        cell_cp.fill = _fill(shade)
        cell_cp.alignment = Alignment(horizontal="center", vertical="center")
        cell_cp.border = border

        # 로그
        buy_str = f"공급가 {buy_price:,}원" if buy_price > 0 else "공급가 미입력"
        n_str = f"네이버 {n_margin:+,}원({n_pct:+.1f}%)" if naver_price > 0 and buy_price > 0 else f"네이버 {naver_price:,}원" if naver_price > 0 else "네이버 -"
        c_str = f"쿠팡 {c_margin:+,}원({c_pct:+.1f}%)" if coupang_price > 0 and buy_price > 0 else f"쿠팡 {coupang_price:,}원" if coupang_price > 0 else "쿠팡 -"
        print(f"   [{i+1}/{len(product_list)}] {name}: {buy_str} → {n_str} | {c_str}")

        results.append({
            "product_name": name,
            "buy_price": buy_price,
            "naver_price": naver_price,
            "coupang_price": coupang_price,
            "naver_margin": n_margin,
            "naver_margin_pct": n_pct,
            "coupang_margin": c_margin,
            "coupang_margin_pct": c_pct,
        })

    wb.save(excel_path)

    # 요약
    profitable_n = sum(1 for r in results if r["naver_margin"] > 0 and r["naver_price"] > 0)
    profitable_c = sum(1 for r in results if r["coupang_margin"] > 0 and r["coupang_price"] > 0)
    total = len(results)

    print(f"\n📊 마진 분석 결과:")
    print(f"   네이버 수익 가능: {profitable_n}/{total}개 (수수료 {NAVER_FEE*100:.0f}%)")
    print(f"   쿠팡 수익 가능:   {profitable_c}/{total}개 (수수료 {COUPANG_FEE*100:.0f}%)")
    print(f"\n✅ 마진 계산 완료 → {excel_path}")

    return results


if __name__ == "__main__":
    print("이 모듈은 main.py를 통해 실행해 주세요.")
