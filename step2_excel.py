"""
STEP 2 — 엑셀 자동 생성 (제품 정보 + 썸네일 이미지)

📊 포함 정보:
  - 기본: No, 이미지, 제품명, 브랜드, 카테고리, 규격, 가격, 특징, 사진수
  - 상세: 제조사, 원산지, 주요성분, 용법, 인증정보, 바코드
  - 마케팅: 포장형태, 타겟, 셀링포인트
  - 네이버: 최저가, 쇼핑몰, 비고 (Step 3에서 채움)
"""
import os
from datetime import datetime
from pathlib import Path
from PIL import Image, ImageOps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from config import THUMBNAIL_SIZE

# ─── 스타일 상수 ────────────────────────────────────────────
NAVY    = "1B2A4A"
TEAL    = "17A589"
WHITE   = "FFFFFF"
LGRAY   = "F7F9FC"
MGRAY   = "EEF3FB"
ACCENT  = "2E6DB4"
RED     = "C00000"
GREEN   = "0D7C3F"
ORANGE  = "D35400"

def _fill(c):
    return PatternFill("solid", start_color=c, fgColor=c)

def _border(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

# ─── 컬럼 정의 ──────────────────────────────────────────────
COLUMNS = [
    # (헤더명, 키, 너비, 정렬, 그룹색)
    ("No.",           None,              5,  CENTER, NAVY),
    ("제품 이미지",    None,             18,  CENTER, NAVY),
    ("제품명",        "product_name",    28,  LEFT,   NAVY),
    ("브랜드/제조사",  "brand",           16,  CENTER, NAVY),
    ("카테고리",      "category",        14,  CENTER, NAVY),
    ("규격/용량",     "spec",            16,  CENTER, NAVY),
    ("확인 가격",     "price_tag",       11,  CENTER, NAVY),
    ("제품 특징",     "feature",         38,  LEFT,   NAVY),
    ("사진수",        None,               7,  CENTER, NAVY),
    # 상세정보 그룹
    ("제조사",        "manufacturer",    14,  CENTER, TEAL),
    ("원산지",        "origin",          10,  CENTER, TEAL),
    ("주요 성분",     "ingredients",     30,  LEFT,   TEAL),
    ("용법/용량",     "usage",           24,  LEFT,   TEAL),
    ("인증 정보",     "certification",   16,  CENTER, TEAL),
    ("바코드",        "barcode",         16,  CENTER, TEAL),
    ("유통기한",      "expiry_info",     14,  CENTER, TEAL),
    # 마케팅 그룹
    ("포장 형태",     "package_type",    10,  CENTER, ORANGE),
    ("타겟 소비자",   "target_audience", 12,  CENTER, ORANGE),
    ("셀링포인트",    "selling_point",   34,  LEFT,   ORANGE),
    # 네이버 가격 (Step 3에서 채움)
    ("온라인 최저가", None,              14,  CENTER, GREEN),
    ("확인 쇼핑몰",  None,              16,  CENTER, GREEN),
    ("비고",         None,              26,  LEFT,   GREEN),
]


# ─── 썸네일 생성 ────────────────────────────────────────────
def make_thumbnail(src_path, dst_path, size=THUMBNAIL_SIZE):
    img = Image.open(src_path)
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass
    img = img.convert("RGB")
    img.thumbnail((size, size), Image.LANCZOS)
    img.save(dst_path, "JPEG", quality=85)
    return dst_path

def pick_representative_image(image_files):
    if not image_files:
        return None
    mid = len(image_files) // 3
    return image_files[min(mid, len(image_files)-1)]


# ─── 엑셀 생성 ──────────────────────────────────────────────
def run(products, output_path):
    print("=" * 60)
    print("  STEP 2: 엑셀 자동 생성")
    print("=" * 60)

    thumb_dir = os.path.join(os.path.dirname(output_path), ".thumbnails")
    os.makedirs(thumb_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "매입 제품 목록"
    ws.sheet_view.showGridLines = False

    num_cols = len(COLUMNS)
    last_col_letter = get_column_letter(num_cols)

    # ── 열 너비 설정 ──
    for i, (_, _, width, _, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Row 1: 메인 타이틀 ──
    ws.row_dimensions[1].height = 44
    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = "📦 토요일 매입건 — 제품 정리 목록"
    c.font = Font(name="맑은 고딕", bold=True, size=16, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = CENTER

    # ── Row 2: 서브 타이틀 (요약 정보) ──
    ws.row_dimensions[2].height = 22
    ws.merge_cells(f"A2:{last_col_letter}2")
    c = ws["A2"]
    now = datetime.now()
    total_img = sum(len(p.get("files", [])) for p in products.values())
    c.value = (
        f"작성일: {now.strftime('%Y-%m-%d (%a)')}  |  "
        f"총 제품: {len(products)}종  |  "
        f"총 이미지: {total_img}장  |  "
        f"온라인 최저가 출처: 네이버 쇼핑 API"
    )
    c.font = Font(name="맑은 고딕", size=9, color="555555")
    c.fill = _fill(MGRAY)
    c.alignment = CENTER

    # ── Row 3: 그룹 헤더 (기본정보 / 상세정보 / 마케팅 / 가격비교) ──
    ws.row_dimensions[3].height = 22
    # 기본정보: A~I (9개)
    ws.merge_cells("A3:I3")
    gc = ws["A3"]
    gc.value = "📋 기본 정보"
    gc.font = Font(name="맑은 고딕", bold=True, size=10, color=WHITE)
    gc.fill = _fill(NAVY)
    gc.alignment = CENTER

    # 상세정보: J~P (7개)
    ws.merge_cells("J3:P3")
    gc = ws["J3"]
    gc.value = "🔬 상세 정보"
    gc.font = Font(name="맑은 고딕", bold=True, size=10, color=WHITE)
    gc.fill = _fill(TEAL)
    gc.alignment = CENTER

    # 마케팅: Q~S (3개)
    ws.merge_cells("Q3:S3")
    gc = ws["Q3"]
    gc.value = "🎯 마케팅"
    gc.font = Font(name="맑은 고딕", bold=True, size=10, color=WHITE)
    gc.fill = _fill(ORANGE)
    gc.alignment = CENTER

    # 가격비교: T~V (3개)
    ws.merge_cells("T3:V3")
    gc = ws["T3"]
    gc.value = "💰 온라인 가격 비교"
    gc.font = Font(name="맑은 고딕", bold=True, size=10, color=WHITE)
    gc.fill = _fill(GREEN)
    gc.alignment = CENTER

    # row3 나머지 셀 채우기
    for i in range(1, num_cols + 1):
        c = ws.cell(row=3, column=i)
        if not c.value:
            c.fill = _fill(NAVY)

    # ── Row 4: 컬럼 헤더 ──
    ws.row_dimensions[4].height = 30
    med_border = _border("thin", "999999")
    for i, (header, _, _, _, group_color) in enumerate(COLUMNS, 1):
        c = ws.cell(row=4, column=i)
        c.value = header
        c.font = Font(name="맑은 고딕", bold=True, size=9, color=WHITE)
        c.fill = _fill(group_color)
        c.alignment = CENTER
        c.border = med_border

    # ── 데이터 행 ──
    thin_b = _border()
    product_list = []

    for idx, (name, data) in enumerate(sorted(products.items()), 1):
        row = idx + 4  # 데이터는 5행부터
        info = data.get("info", {})
        files = data.get("files", [])
        shade = LGRAY if idx % 2 == 1 else WHITE
        ws.row_dimensions[row].height = 110

        def wc(col, val, align=CENTER, fnt=None, color=None):
            c = ws.cell(row=row, column=col)
            c.value = val
            c.font = fnt or Font(name="맑은 고딕", size=9, color=color or "333333")
            c.fill = _fill(shade)
            c.alignment = align
            c.border = thin_b

        # 기본 정보
        wc(1, idx)
        wc(2, "")  # 이미지 placeholder
        wc(3, info.get("product_name", name), LEFT,
           Font(name="맑은 고딕", bold=True, size=10, color="1a1a1a"))
        wc(4, info.get("brand", ""))
        wc(5, info.get("category", ""))
        wc(6, info.get("spec", ""))

        # 가격은 빨간 볼드
        price_tag = info.get("price_tag", "")
        wc(7, price_tag, CENTER,
           Font(name="맑은 고딕", bold=True, size=11, color=RED))

        wc(8, info.get("feature", ""), LEFT)
        wc(9, len(files))

        # 상세 정보
        wc(10, info.get("manufacturer", ""))
        wc(11, info.get("origin", ""))
        wc(12, info.get("ingredients", ""), LEFT, Font(name="맑은 고딕", size=8))
        wc(13, info.get("usage", ""), LEFT, Font(name="맑은 고딕", size=8))
        wc(14, info.get("certification", ""))
        wc(15, info.get("barcode", ""))
        wc(16, info.get("expiry_info", ""))

        # 마케팅
        wc(17, info.get("package_type", ""))
        wc(18, info.get("target_audience", ""))
        wc(19, info.get("selling_point", ""), LEFT, Font(name="맑은 고딕", size=8))

        # 네이버 가격 (Step 3에서 채움)
        wc(20, "", CENTER, Font(name="맑은 고딕", bold=True, size=11, color=RED))
        wc(21, "")
        wc(22, "", LEFT)

        # 썸네일 삽입
        rep_img = pick_representative_image(files)
        if rep_img:
            thumb_path = os.path.join(thumb_dir, f"thumb_{idx:02d}.jpg")
            try:
                make_thumbnail(rep_img, thumb_path)
                xl_img = XLImage(thumb_path)
                xl_img.width, xl_img.height = THUMBNAIL_SIZE, THUMBNAIL_SIZE
                ws.add_image(xl_img, f"B{row}")
            except Exception as e:
                print(f"   ⚠️  썸네일 오류 ({name}): {e}")

        product_list.append({
            "row": row,
            "name": info.get("product_name", name),
            "brand": info.get("brand", ""),
            "folder": data.get("folder_path", ""),
            "files": files,
        })
        print(f"   [{idx}/{len(products)}] {info.get('product_name', name)} ✓")

    # ── 푸터 ──
    footer_row = len(products) + 5
    ws.row_dimensions[footer_row].height = 24
    ws.merge_cells(f"A{footer_row}:{last_col_letter}{footer_row}")
    c = ws.cell(row=footer_row, column=1)
    c.value = f"총 {len(products)}개 제품  |  {total_img}장 이미지  |  생성: {now.strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(name="맑은 고딕", size=9, italic=True, color="888888")
    c.fill = _fill(MGRAY)
    c.alignment = CENTER

    # ── 시트 설정 ──
    ws.freeze_panes = "C5"  # A~B열, 1~4행 고정
    ws.auto_filter.ref = f"A4:{last_col_letter}{footer_row - 1}"

    wb.save(output_path)
    print(f"\n✅ 엑셀 저장 완료 → {output_path}")
    print(f"   - {len(products)}개 제품, {total_img}장 이미지")
    print(f"   - 컬럼: {num_cols}개 (기본9 + 상세7 + 마케팅3 + 가격3)")

    return product_list, output_path

if __name__ == "__main__":
    print("이 모듈은 main.py를 통해 실행해 주세요.")
