"""
STEP 3-2 — 쿠팡 판매 확인 + 최저가 비교 (고도화)

1차: curl_cffi로 쿠팡 검색 직접 크롤링 (Akamai 우회)
2차: 네이버 쇼핑 API에서 mallName='쿠팡' 필터링 (폴백)

curl_cffi 설치: pip install curl_cffi
"""
import os, json, time, re
import urllib.request, urllib.parse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from config import NAVER_CLIENT_ID, NAVER_CLIENT_SECRET, NAVER_DISPLAY
except ImportError:
    NAVER_CLIENT_ID = ""
    NAVER_CLIENT_SECRET = ""
    NAVER_DISPLAY = 20

try:
    from curl_cffi import requests as curl_req
    HAS_CURL = True
except ImportError:
    HAS_CURL = False

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

# ─── 스타일 상수
ROCKET_GREEN = "27AE60"
GRAY_TEXT    = "AAAAAA"

def _fill(c):
    return PatternFill("solid", start_color=c, fgColor=c)

def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


# ─── 방법 1: curl_cffi로 쿠팡 직접 크롤링 ──────────────────────
def search_coupang_direct(query):
    """
    curl_cffi + Chrome impersonation으로 쿠팡 검색 직접 크롤링.
    Akamai CDN 봇 탐지를 실제 브라우저 TLS로 우회.
    """
    if not HAS_CURL or not HAS_BS4:
        return []

    encoded = urllib.parse.quote(query)
    url = (
        f"https://www.coupang.com/np/search"
        f"?q={encoded}&channel=user&sorter=scoreDesc&listSize=36"
    )
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer": "https://www.coupang.com/",
        "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    try:
        resp = curl_req.get(url, headers=headers, impersonate="chrome120", timeout=15)
        if resp.status_code != 200:
            print(f"   ⚠️  쿠팡 직접 접근 실패 (HTTP {resp.status_code})")
            return []

        soup = BeautifulSoup(resp.text, "html.parser")
        items = []

        # 검색 결과 파싱
        for li in soup.select("li.search-product")[:20]:
            try:
                name_el = li.select_one("div.name") or li.select_one("span.name")
                price_el = li.select_one("strong.price-value") or li.select_one("em.price-value")
                link_el = li.select_one("a.search-product-link") or li.select_one("a[href*='/vp/products/']")

                if not name_el:
                    continue

                title = name_el.get_text(strip=True)
                price_str = price_el.get_text(strip=True).replace(",", "") if price_el else "0"
                try:
                    price = int(re.sub(r"[^0-9]", "", price_str))
                except ValueError:
                    price = 0

                href = link_el.get("href", "") if link_el else ""
                link = ("https://www.coupang.com" + href) if href.startswith("/") else href

                # 로켓배송 여부
                is_rocket = bool(
                    li.select_one("span.badge-rocket") or
                    li.select_one("img[alt*='로켓']") or
                    li.select_one(".rocket-badge") or
                    "rocket" in li.get("class", [])
                )

                if price > 0:
                    items.append({
                        "title": title,
                        "price": price,
                        "link": link,
                        "mall": "쿠팡",
                        "is_rocket": is_rocket,
                    })
            except Exception:
                continue

        return items

    except Exception as e:
        print(f"   ⚠️  쿠팡 직접 크롤링 오류: {e}")
        return []


# ─── 방법 2: 네이버 쇼핑 API에서 쿠팡 필터링 (폴백) ─────────────
def search_naver_for_coupang(query, display=100):
    """
    네이버 쇼핑 API로 검색 → mallName/link에 '쿠팡' 포함 항목 필터링.
    display=100으로 최대한 많이 검색.
    """
    encoded = urllib.parse.quote(query)
    url = f"https://openapi.naver.com/v1/search/shop.json?query={encoded}&display={display}&sort=asc"

    req = urllib.request.Request(url)
    req.add_header("X-Naver-Client-Id", NAVER_CLIENT_ID)
    req.add_header("X-Naver-Client-Secret", NAVER_CLIENT_SECRET)

    try:
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read().decode("utf-8"))
        items = data.get("items", [])
    except Exception as e:
        print(f"   ⚠️  네이버 검색 오류: {e}")
        return {"all_items": [], "coupang_items": []}

    all_items = []
    coupang_items = []

    for item in items:
        title = re.sub(r"<[^>]+>", "", item.get("title", ""))
        mall = item.get("mallName", "")
        link = item.get("link", "")
        try:
            price = int(item.get("lprice", 0))
        except (ValueError, TypeError):
            price = 0

        parsed = {
            "title": title,
            "mall": mall,
            "price": price,
            "link": link,
            "image": item.get("image", ""),
        }
        all_items.append(parsed)

        # 쿠팡 판매 여부 확인 (대소문자, 로켓 포함)
        is_coupang = (
            "쿠팡" in mall
            or "coupang" in mall.lower()
            or "coupang" in link.lower()
            or "로켓" in mall
        )
        if is_coupang and price > 0:
            coupang_items.append(parsed)

    return {"all_items": all_items, "coupang_items": coupang_items}


def _build_coupang_status(items, source):
    """쿠팡 아이템 리스트 → 상태 딕셔너리"""
    if not items:
        return None
    lowest = min(items, key=lambda x: x["price"])
    return {
        "on_coupang": True,
        "coupang_price": lowest["price"],
        "coupang_price_str": f"{lowest['price']:,}원",
        "coupang_count": len(items),
        "coupang_title": lowest["title"],
        "coupang_link": lowest.get("link", ""),
        "source": source,
        "status_text": f"쿠팡 판매중 ({len(items)}건, {source})",
    }


def find_coupang(query, brand=""):
    """
    1차: curl_cffi 직접 크롤링
    2차: 네이버 API 폴백
    검색어도 여러 개 시도 (브랜드+제품명, 제품명만)
    """
    queries = []
    if brand:
        queries.append(f"{brand} {query}".strip())
    queries.append(query)
    # 제품명이 길면 앞 부분만 잘라서도 시도
    if len(query) > 10:
        queries.append(query[:10].strip())

    # 1차: 직접 크롤링
    if HAS_CURL and HAS_BS4:
        for q in queries:
            items = search_coupang_direct(q)
            if items:
                status = _build_coupang_status(items, "직접크롤링")
                if status:
                    return status
        # 직접 크롤링 결과 없어도 API 시도

    # 2차: 네이버 API
    if NAVER_CLIENT_ID and NAVER_CLIENT_SECRET:
        all_coupang = []
        for q in queries:
            result = search_naver_for_coupang(q, display=100)
            all_coupang.extend(result.get("coupang_items", []))
            if all_coupang:
                break  # 결과 있으면 더 시도 안 함

        if all_coupang:
            # 중복 제거 (가격 기준)
            seen = set()
            unique = []
            for item in all_coupang:
                key = (item["price"], item["title"][:20])
                if key not in seen:
                    seen.add(key)
                    unique.append(item)
            return _build_coupang_status(unique, "네이버API")

    # 둘 다 실패
    return {
        "on_coupang": False,
        "coupang_price": 0,
        "coupang_price_str": "",
        "coupang_count": 0,
        "coupang_title": "",
        "coupang_link": "",
        "source": "",
        "status_text": "쿠팡 미판매",
    }


# ─── 엑셀 업데이트 ──────────────────────────────────────────
def update_excel(excel_path, product_list):
    wb = load_workbook(excel_path)
    ws = wb.active
    border = _border()
    results = []
    total = len(product_list)

    method = "curl_cffi(직접)" if HAS_CURL else "네이버API"
    print(f"\n   조회 방식: {method}")

    for i, product in enumerate(product_list):
        row = product["row"]
        name = product["name"]
        brand = product.get("brand", "")

        print(f"   [{i+1}/{total}] 검색: {brand+' '+name if brand else name}")

        status = find_coupang(name, brand)

        shade = "F5F5F5" if i % 2 == 0 else "FFFFFF"

        col_cprice = 23  # W: 쿠팡 최저가
        col_status = 24  # X: 쿠팡 판매 상태

        if status["on_coupang"]:
            c_price = ws.cell(row=row, column=col_cprice)
            c_price.value = status["coupang_price_str"]
            c_price.font = Font(name="맑은 고딕", bold=True, size=10, color="C00000")
            c_price.fill = _fill(shade)
            c_price.alignment = Alignment(horizontal="center", vertical="center")
            c_price.border = border

            c_status = ws.cell(row=row, column=col_status)
            c_status.value = f"판매중 ({status['coupang_count']}건)"
            c_status.font = Font(name="맑은 고딕", bold=True, size=10, color=ROCKET_GREEN)
            c_status.fill = _fill(shade)
            c_status.alignment = Alignment(horizontal="center", vertical="center")
            c_status.border = border

            src = status.get("source", "")
            print(f"      → 쿠팡 {status['coupang_price_str']} ({status['coupang_count']}건) [{src}]")
        else:
            c_price = ws.cell(row=row, column=col_cprice)
            c_price.value = "-"
            c_price.font = Font(name="맑은 고딕", size=9, color=GRAY_TEXT)
            c_price.fill = _fill(shade)
            c_price.alignment = Alignment(horizontal="center", vertical="center")
            c_price.border = border

            c_status = ws.cell(row=row, column=col_status)
            c_status.value = "미판매"
            c_status.font = Font(name="맑은 고딕", size=9, color=GRAY_TEXT)
            c_status.fill = _fill(shade)
            c_status.alignment = Alignment(horizontal="center", vertical="center")
            c_status.border = border

            print(f"      → 쿠팡 미판매")

        results.append({
            "product_name": name,
            "on_coupang": status["on_coupang"],
            "coupang_price": status["coupang_price"],
            "coupang_price_str": status["coupang_price_str"],
            "coupang_count": status["coupang_count"],
            "coupang_link": status["coupang_link"],
            "status_text": status["status_text"],
        })

        time.sleep(0.5)

    wb.save(excel_path)
    print(f"\n✅ 엑셀 업데이트 완료 → {excel_path}")
    return results


# ─── 메인 실행 ─────────────────────────────────────────────
def run(product_list, excel_path):
    print("=" * 60)
    print("  STEP 3-2: 쿠팡 판매 확인 (고도화)")
    print("=" * 60)

    if not product_list:
        print("\n⚠️  제품 목록이 없습니다.")
        return []

    if not excel_path or not os.path.exists(excel_path):
        print(f"\n⚠️  엑셀 파일을 찾을 수 없습니다: {excel_path}")
        return []

    # 방식 안내
    if HAS_CURL and HAS_BS4:
        print("\n🚀 curl_cffi 직접 크롤링 모드 (Akamai 우회)")
        print("   → 쿠팡 검색 페이지를 직접 크롤링합니다")
    elif NAVER_CLIENT_ID and NAVER_CLIENT_SECRET:
        print("\n📡 네이버 쇼핑 API 폴백 모드")
        print("   ⚠️  curl_cffi 미설치 → pip install curl_cffi beautifulsoup4")
    else:
        print("\n⚠️  curl_cffi도 없고 네이버 API 키도 없습니다.")
        return []

    print(f"\n🔍 {len(product_list)}개 제품의 쿠팡 판매 여부를 확인합니다...\n")
    results = update_excel(excel_path, product_list)

    on_coupang = sum(1 for r in results if r.get("on_coupang"))
    print(f"\n📊 쿠팡 확인 결과:")
    print(f"   - 쿠팡 판매중: {on_coupang}/{len(results)}개")
    print(f"   - 쿠팡 미판매: {len(results) - on_coupang}/{len(results)}개")

    for r in results:
        icon = "🟢" if r["on_coupang"] else "⚪"
        price = r.get("coupang_price_str", "-") or "-"
        print(f"   {icon} {r['product_name']}: {price} ({r['status_text']})")

    return results


if __name__ == "__main__":
    print("이 모듈은 main.py를 통해 실행해 주세요.")
